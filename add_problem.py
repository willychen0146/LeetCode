import glob
import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


# Function to convert the leetcode naming format
def convert_to_title_format(name):
    # Split the name into words using "-" as delimiter
    words = name.split("-")
    
    # Capitalize the first letter of each word
    title_words = [word.capitalize() for word in words]
    
    # Join the words back together with space as separator
    title_formatted = " ".join(title_words)
    
    return title_formatted

# Prompt 
def prompt(message):
    return input(message)

# Convert hexadecimal color code to RGB
def hex_to_rgb(hex_color):
    return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

def find_java_files(folder_path):
    java_files = glob.glob(os.path.join(folder_path, '**', '*.java'), recursive=True)
    return java_files

# easy color
easy = '98fb98'
rgb_easy = hex_to_rgb(easy)
fill_easy = PatternFill(start_color=easy, end_color=easy, fill_type='solid')

# mediem
medium = 'ffd700'
rgb_medium = hex_to_rgb(medium)
fill_medium = PatternFill(start_color=medium, end_color=medium, fill_type='solid')

# hard
hard = 'f08080'
rgb_hard = hex_to_rgb(hard)
fill_hard = PatternFill(start_color=hard, end_color=hard, fill_type='solid')

# Prompt for the LeetCode problem URL
url = prompt('Please input the LeetCode problem URL: ')

# Extract problem name from the URL
match = re.match(r'https://leetcode.com/problems/(.*)', url)
if match:
    problem_name = match.group(1).split('/')[0]

    # Attempt to create directory for the problem
    try:
        os.mkdir(problem_name)
        print(f"{problem_name} directory created.")
    except FileExistsError:
        print(f"{problem_name} directory already exists.")
    except Exception as e:
        print(f"Error creating {problem_name} directory:", e)

    # Check if the file already exists
    solution_file_path = f'./{problem_name}/Solution.java'
    directory_path = f'./{problem_name}/'
    java_files = find_java_files(directory_path)
    if not java_files:
        try:
            # Creates a new java file
            with open(solution_file_path, 'w') as fp:
                pass
            print(f"Solution.java created.")
        except Exception as e:
            print(f"Error creating Solution.java", e)
    else:
        print(f"Solution.java already exists.")

    # create excel
    try:
        # Check if Excel file already exists
        if not os.path.exists('note.xlsx'):
            wb = Workbook()
            wb.save('note.xlsx')
            print("Excel file created.")
        else:
            print("Excel file already exists. Skipping creation.")

        # Load existing workbook
        wb = load_workbook('note.xlsx')
        sheet = wb.active
        
        # Add labels with bold font if not already present
        if sheet['A1'].value != "Category":
            bold_font = Font(bold=True)
            sheet['A1'] = "Category"
            sheet['A1'].font = bold_font
            sheet['B1'] = "Name"
            sheet['B1'].font = bold_font
            sheet['C1'] = "Note"
            sheet['C1'].font = bold_font
            sheet['D1'] = "Link"
            sheet['D1'].font = bold_font

        # Search for the category in existing rows
        category = prompt('Please input a Category: ')
        category_row = None
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1):
            if row[0].value == category:
                category_row = row[0].row
                break

        if category_row:
            # Find the next available row
            next_row = len(sheet['A']) + 1
            
            # Insert new row below the category and shift down other rows
            sheet.insert_rows(category_row + 1)

            # Write data to the new row
            sheet[f'A{category_row + 1}'] = category
            sheet[f'B{category_row + 1}'] = problem_name
            level = prompt('Please input a problem level (e/m/h): ')
            if level == 'e':
                sheet[f'B{category_row + 1}'].fill = fill_easy
            if level == 'm':
                sheet[f'B{category_row + 1}'].fill = fill_medium
            if level == 'h':
                sheet[f'B{category_row + 1}'].fill = fill_hard
            note = prompt('Please input a note: ')
            sheet[f'C{category_row + 1}'] = note
            sheet[f'D{category_row + 1}'] = url

        else:
            # Find the next available row
            next_row = len(sheet['A']) + 1

            # Write data to the next available row
            sheet[f'A{next_row}'] = category
            sheet[f'B{next_row}'] = problem_name
            level = prompt('Please input a problem level: (e/m/h): ')
            if level == 'e':
                sheet[f'B{next_row}'].fill = fill_easy
            if level == 'm':
                sheet[f'B{next_row}'].fill = fill_medium
            if level == 'h':
                sheet[f'B{next_row}'].fill = fill_hard
            note = prompt('Please input a note: ')
            sheet[f'C{next_row}'] = note
            sheet[f'D{next_row}'] = url

        # Save changes to the workbook
        wb.save('note.xlsx')
        print("Data written to Excel file.")

    except Exception as e:
        print("Error:", e)

else:
    print("Invalid URL format.")
